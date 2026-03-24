#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel报告生成工具 - AI投资Agent
生成标的深度分析Excel报告
"""

import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    sys.exit(1)


def create_position_analysis_excel(data, output_path):
    """
    创建标的深度分析Excel

    data格式: [
        {
            "板块": "科技",
            "标的": "腾讯控股",
            "代码": "00700.HK",
            "持仓比例": "5%",
            "成本价": "350",
            "现价": "380",
            "PE": "18.5",
            "历史分位": "35%",
            "基本面": "...",
            "估值分析": "...",
            "护城河": "...",
            "驱动因素": "...",
            "风险因素": "...",
            "趋势买卖点": "...",
            "长期温度": "低估",
            "短期温度": "合理",
            "综合评级": "★★★★☆",
            "操作建议": "..."
        },
        ...
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "标的深度分析"

    # 定义列
    columns = [
        ("板块", 8),
        ("标的", 12),
        ("代码", 12),
        ("持仓比例", 10),
        ("成本价", 10),
        ("现价", 10),
        ("PE", 8),
        ("历史分位", 10),
        ("基本面", 40),
        ("估值分析", 40),
        ("护城河", 30),
        ("驱动因素", 35),
        ("风险因素", 35),
        ("趋势买卖点", 30),
        ("长期温度", 10),
        ("短期温度", 10),
        ("综合评级", 12),
        ("操作建议", 50)
    ]

    # 样式定义
    header_font = Font(name='楷体', bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='楷体', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 温度颜色映射
    temp_colors = {
        "严重低估": "00B050",  # 深绿
        "低估": "92D050",      # 浅绿
        "合理": "FFFF00",      # 黄色
        "高估": "FFC000",      # 橙色
        "严重高估": "FF0000"   # 红色
    }

    # 写入表头
    for col, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(name='楷体', bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # 写入数据
    for row_idx, item in enumerate(data, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            value = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 温度列特殊着色
            if key in ["长期温度", "短期温度"] and value in temp_colors:
                cell.fill = PatternFill(
                    start_color=temp_colors[value],
                    end_color=temp_colors[value],
                    fill_type="solid"
                )

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置行高
    ws.row_dimensions[1].height = 25
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 80

    wb.save(output_path)
    print(f"已保存: {output_path}")


def create_temperature_excel(data, output_path):
    """
    创建温度计Excel

    data格式: [
        {
            "市场": "A股",
            "板块": "宽基",
            "指数": "沪深300",
            "代码": "000300",
            "当日涨跌": "+1.2%",
            "今年涨跌": "+5.3%",
            "PE": "12.5",
            "温度": "35"
        },
        ...
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "指数温度计"

    columns = [
        ("市场", 8),
        ("板块", 10),
        ("指数", 15),
        ("代码", 10),
        ("当日涨跌", 10),
        ("今年涨跌", 10),
        ("PE(TTM)", 10),
        ("温度", 8),
        ("操作建议", 15)
    ]

    # 样式
    header_font = Font(name='楷体', bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell_font = Font(name='楷体', size=10)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # 温度颜色函数
    def get_temp_color(temp):
        try:
            t = float(temp)
            if t <= 20:
                return "00B050"  # 深绿
            elif t <= 30:
                return "92D050"  # 浅绿
            elif t <= 50:
                return "FFFF00"  # 黄色
            elif t <= 70:
                return "FFC000"  # 橙色
            else:
                return "FF0000"  # 红色
        except:
            return "FFFFFF"

    # 温度对应操作建议
    def get_suggestion(temp):
        try:
            t = float(temp)
            if t <= 20:
                return "重仓买入"
            elif t <= 30:
                return "分批买入"
            elif t <= 50:
                return "持有观望"
            elif t <= 70:
                return "谨慎持有"
            else:
                return "分批卖出"
        except:
            return "-"

    # 写入数据
    for row_idx, item in enumerate(data, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            if key == "操作建议":
                value = get_suggestion(item.get("温度", ""))
            else:
                value = item.get(key, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # 温度列着色
            if key == "温度":
                cell.fill = PatternFill(
                    start_color=get_temp_color(value),
                    end_color=get_temp_color(value),
                    fill_type="solid"
                )

    # 冻结首行
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"已保存: {output_path}")


# 示例用法
if __name__ == '__main__':
    # 示例数据
    sample_data = [
        {
            "板块": "科技",
            "标的": "腾讯控股",
            "代码": "00700.HK",
            "持仓比例": "5%",
            "成本价": "350",
            "现价": "380",
            "PE": "18.5",
            "历史分位": "35%",
            "基本面": "2024年营收6740亿，净利润1900亿，同比增长35%",
            "估值分析": "PE 18.5x，历史分位35%，低于5年中位数",
            "护城河": "社交+支付+游戏生态，品牌/技术/网络效应强",
            "驱动因素": "1.游戏版号恢复 2.视频号商业化 3.AI能力提升",
            "风险因素": "1.监管政策 2.游戏行业竞争 3.宏观经济放缓",
            "趋势买卖点": "支撑位350，压力位420，建议买入区间350-370",
            "长期温度": "低估",
            "短期温度": "合理",
            "综合评级": "★★★★☆",
            "操作建议": "当前估值合理偏低，可逢低分批建仓"
        }
    ]

    output_path = f"{datetime.now().strftime('%Y-%m-%d')}_标的深度分析_示例.xlsx"
    create_position_analysis_excel(sample_data, output_path)
